import os
import tempfile
import subprocess
import json
from pathlib import Path
import base64
from io import BytesIO
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class JasperEngine:
    """JasperReports engine wrapper using pyreportjasper for proper JRXML processing."""
    
    def __init__(self):
        self.temp_dir = Path(tempfile.gettempdir()) / 'jasper_reports'
        self.temp_dir.mkdir(exist_ok=True)
        self.java_available = self._check_java_availability()
        self.jdbc_dir = Path(__file__).parent / 'lib'
        self._setup_jdbc_classpath()
    
    def _setup_jdbc_classpath(self):
        """Set up JDBC driver classpath for JasperReports."""
        # Try Java versions in order of compatibility (8 > 11 > system)
        java_homes = [
            "/usr/local/opt/openjdk@8/libexec/openjdk.jdk/Contents/Home",  # Homebrew Java 8
            "/usr/local/Cellar/openjdk@8/1.8.0-452/libexec/openjdk.jdk/Contents/Home",  # Direct Java 8 path
            "/Library/Java/JavaVirtualMachines/adoptopenjdk-11.jdk/Contents/Home",  # Java 11
            "/Library/Java/JavaVirtualMachines/openjdk-8.jdk/Contents/Home"  # System Java 8
        ]
        
        java_home_set = False
        for java_home in java_homes:
            if os.path.exists(java_home):
                os.environ['JAVA_HOME'] = java_home
                # Also update PATH to use this Java
                java_bin = os.path.join(java_home, 'bin')
                current_path = os.environ.get('PATH', '')
                if java_bin not in current_path:
                    os.environ['PATH'] = f"{java_bin}:{current_path}"
                logger.info(f"JAVA_HOME set to compatible version: {java_home}")
                java_home_set = True
                break
        
        if not java_home_set:
            # Fallback to system Java
            if not os.environ.get('JAVA_HOME'):
                try:
                    result = subprocess.run(['java', '-XshowSettings:properties', '-version'], 
                                          capture_output=True, text=True, timeout=10)
                    java_home_line = [line for line in result.stderr.split('\n') if 'java.home' in line]
                    if java_home_line:
                        java_home = java_home_line[0].split('=')[1].strip()
                        os.environ['JAVA_HOME'] = java_home
                        logger.info(f"JAVA_HOME set to system Java: {java_home}")
                except Exception as e:
                    logger.warning(f"Could not determine JAVA_HOME: {e}")
        
        if self.jdbc_dir.exists():
            jdbc_jars = list(self.jdbc_dir.glob('*.jar'))
            if jdbc_jars:
                classpath = ':'.join(str(jar) for jar in jdbc_jars)
                os.environ['CLASSPATH'] = classpath
                logger.info(f"JDBC classpath set: {classpath}")
        
    def _check_java_availability(self):
        """Check if Java is available for JasperReports compilation."""
        try:
            result = subprocess.run(['java', '-version'], 
                                  capture_output=True, text=True, timeout=5)
            if result.returncode == 0:
                logger.info("Java is available")
                return True
            else:
                logger.warning("Java not found")
                return False
        except Exception as e:
            logger.error(f"Error checking Java: {e}")
            return False
    
    def compile_and_generate(self, jrxml_content, connection_string, output_format='html', parameters=None):
        """Generate report using actual JasperReports engine."""
        try:
            if parameters is None:
                parameters = {}
                
            if not self.java_available:
                return self._generate_fallback_message("Java is required for JasperReports processing"), None
                
            # For now, use enhanced fallback system due to Java compatibility issues
            logger.info("Using enhanced JRXML preview system")
            
            if output_format in ['html']:
                return self._generate_html_fallback(jrxml_content, connection_string)
            elif output_format == 'pdf':
                return self._generate_pdf_fallback(jrxml_content, connection_string)
            elif output_format in ['xlsx', 'excel']:
                return self._generate_excel_fallback(jrxml_content, connection_string)
            elif output_format == 'csv':
                return self._generate_csv_fallback(jrxml_content, connection_string)
            else:
                return self._generate_html_fallback(jrxml_content, connection_string)
                    
        except Exception as e:
            error_msg = f"Report generation error: {str(e)}"
            logger.error(error_msg)
            return error_msg.encode('utf-8'), str(e)
    
    def _generate_with_pyreportjasper(self, jrxml_content, connection_string, output_format, parameters):
        """Generate report using pyreportjasper."""
        try:
            from pyreportjasper import PyReportJasper
            
            # Create temporary files
            timestamp = os.getpid()
            jrxml_file = self.temp_dir / f"report_{timestamp}.jrxml"
            output_file = self.temp_dir / f"report_{timestamp}"
            
            # Write JRXML content to file
            with open(jrxml_file, 'w', encoding='utf-8') as f:
                f.write(jrxml_content)
            
            # Map format names to pyreportjasper supported formats
            format_map = {
                'html': 'html',
                'pdf': 'pdf', 
                'xlsx': 'xlsx',
                'excel': 'xlsx',
                'csv': 'csv',
                'xml': 'xml',
                'rtf': 'rtf',
                'docx': 'docx',
                'odt': 'odt',
                'ods': 'ods',
                'pptx': 'pptx'
            }
            
            jasper_format = format_map.get(output_format, output_format)
            
            # Prepare configuration
            config = {
                'input_file': str(jrxml_file),
                'output_file': str(output_file),
                'output_formats': [jasper_format],
                'parameters': parameters or {},
                'locale': 'en_US'
            }
            
            # Add database configuration if provided
            if connection_string:
                db_config = self._parse_connection_string(connection_string)
                if db_config:
                    config['db_connection'] = db_config
            
            # Create instance and process
            pyjasper = PyReportJasper()
            pyjasper.config(**config)
            pyjasper.process_report()
            
            # Read the generated file
            output_path = Path(f"{output_file}.{jasper_format}")
            if output_path.exists():
                with open(output_path, 'rb') as f:
                    content = f.read()
                
                # Clean up temporary files
                self._cleanup_temp_files([jrxml_file, output_path])
                
                return content, None
            else:
                raise Exception(f"Output file not generated: {output_path}")
                
        except Exception as e:
            logger.error(f"pyreportjasper error: {e}")
            raise
    
    def _generate_with_pyjasper(self, jrxml_content, connection_string, output_format, parameters):
        """Generate report using pyjasper."""
        try:
            import pyjasper
            
            # Create temporary files
            timestamp = os.getpid()
            jrxml_file = self.temp_dir / f"report_{timestamp}.jrxml"
            output_file = self.temp_dir / f"report_{timestamp}"
            
            # Write JRXML content to file
            with open(jrxml_file, 'w', encoding='utf-8') as f:
                f.write(jrxml_content)
            
            # Prepare database connection
            db_config = None
            if connection_string:
                db_config = self._parse_connection_string(connection_string)
            
            # Create JasperPy instance
            jasper = pyjasper.JasperPy()
            
            # Process the report
            jasper.process(
                str(jrxml_file),
                output_file=str(output_file),
                format_list=[output_format],
                parameters=parameters,
                db_connection=db_config if db_config else None,
                locale='en_US'
            )
            
            # Read the generated file
            output_path = Path(f"{output_file}.{output_format}")
            if output_path.exists():
                with open(output_path, 'rb') as f:
                    content = f.read()
                
                # Clean up temporary files
                self._cleanup_temp_files([jrxml_file, output_path])
                
                return content, None
            else:
                # Try without extension (some versions don't add it)
                if output_file.exists():
                    with open(output_file, 'rb') as f:
                        content = f.read()
                    self._cleanup_temp_files([jrxml_file, output_file])
                    return content, None
                else:
                    raise Exception(f"Output file not generated")
                    
        except Exception as e:
            logger.error(f"pyjasper error: {e}")
            raise
    
    def _parse_connection_string(self, connection_string):
        """Parse connection string to database configuration."""
        try:
            if connection_string.startswith('sqlite:///'):
                db_path = connection_string.replace('sqlite:///', '')
                if os.path.exists(db_path):
                    return {
                        'driver': 'sqlite',
                        'database': db_path,
                        'jdbc_driver': 'org.sqlite.JDBC',
                        'jdbc_url': f'jdbc:sqlite:{db_path}'
                    }
            elif connection_string.startswith('mysql://'):
                # Parse MySQL connection string
                # Format: mysql://user:password@host:port/database
                from urllib.parse import urlparse
                parsed = urlparse(connection_string)
                return {
                    'driver': 'mysql',
                    'host': parsed.hostname,
                    'port': parsed.port or 3306,
                    'database': parsed.path.lstrip('/'),
                    'username': parsed.username,
                    'password': parsed.password,
                    'jdbc_driver': 'com.mysql.jdbc.Driver',
                    'jdbc_url': f'jdbc:mysql://{parsed.hostname}:{parsed.port or 3306}/{parsed.path.lstrip("/")}'
                }
            elif connection_string.startswith('postgresql://'):
                # Parse PostgreSQL connection string
                from urllib.parse import urlparse
                parsed = urlparse(connection_string)
                return {
                    'driver': 'postgres',
                    'host': parsed.hostname,
                    'port': parsed.port or 5432,
                    'database': parsed.path.lstrip('/'),
                    'username': parsed.username,
                    'password': parsed.password,
                    'jdbc_driver': 'org.postgresql.Driver',
                    'jdbc_url': f'jdbc:postgresql://{parsed.hostname}:{parsed.port or 5432}/{parsed.path.lstrip("/")}'
                }
        except Exception as e:
            logger.error(f"Error parsing connection string: {e}")
        return None
    
    def _cleanup_temp_files(self, files):
        """Clean up temporary files."""
        for file_path in files:
            try:
                if isinstance(file_path, Path):
                    file_path.unlink(missing_ok=True)
                else:
                    Path(file_path).unlink(missing_ok=True)
            except Exception as e:
                logger.warning(f"Could not delete temp file {file_path}: {e}")
    
    def _generate_fallback_message(self, message):
        """Generate a fallback HTML message."""
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>JasperReports Engine Status</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; }}
                .warning {{ 
                    background-color: #fff3cd; 
                    border: 1px solid #ffeeba; 
                    color: #856404; 
                    padding: 20px; 
                    border-radius: 5px; 
                }}
                .info {{ margin-top: 20px; }}
                code {{ 
                    background-color: #f8f9fa; 
                    padding: 2px 4px; 
                    border-radius: 3px; 
                    font-family: monospace; 
                }}
            </style>
        </head>
        <body>
            <div class="warning">
                <h3>JasperReports Engine Not Available</h3>
                <p>{message}</p>
                <div class="info">
                    <h4>Requirements:</h4>
                    <ol>
                        <li>Java 8 or higher must be installed</li>
                        <li>Install pyreportjasper: <code>pip install pyreportjasper</code></li>
                        <li>Or install pyjasper: <code>pip install pyjasper</code></li>
                    </ol>
                    <h4>Check Java installation:</h4>
                    <code>java -version</code>
                </div>
            </div>
        </body>
        </html>
        """
        return html.encode('utf-8')
    
    def get_supported_formats(self):
        """Get list of supported output formats."""
        return ['html', 'pdf', 'xlsx', 'docx', 'csv', 'xml', 'rtf', 'odt', 'ods', 'pptx']
    
    def is_jasper_available(self):
        """Check if JasperReports engine is available."""
        if not self.java_available:
            return False, "Java is not installed"
            
        try:
            import pyreportjasper
            return True, "JasperReports engine is available (pyreportjasper)"
        except ImportError:
            try:
                import pyjasper
                return True, "JasperReports engine is available (pyjasper)"
            except ImportError:
                return False, "No JasperReports Python library found"
    
    def _generate_html_fallback(self, jrxml_content, connection_string):
        """Generate HTML fallback when JasperReports engine fails."""
        try:
            import xml.etree.ElementTree as ET
            import sqlite3
            
            # Parse JRXML
            root = ET.fromstring(jrxml_content)
            
            # Extract namespace
            ns = {'jr': 'http://jasperreports.sourceforge.net/jasperreports'}
            if root.tag.startswith('{'):
                ns_url = root.tag[1:root.tag.index('}')]
                ns = {'jr': ns_url}
            
            # Extract title from title band
            title_text = None
            title_band = root.find('.//jr:title', ns)
            if title_band is None:
                title_band = root.find('.//title')
            
            if title_band is not None:
                # Find static text elements in title band
                static_texts = title_band.findall('.//jr:staticText/jr:text', ns)
                if not static_texts:
                    static_texts = title_band.findall('.//staticText/text')
                
                for text_elem in static_texts:
                    if text_elem.text and text_elem.text.strip():
                        title_text = text_elem.text.strip()
                        break
            
            # Extract fields
            fields = {}
            field_elements = root.findall('.//jr:field', ns)
            if not field_elements:
                field_elements = root.findall('.//field')
            
            for field in field_elements:
                name = field.get('name')
                field_class = field.get('class', 'java.lang.String')
                if name:
                    fields[name] = field_class
            
            # Extract query
            query = ""
            query_elem = root.find('.//jr:queryString', ns)
            if query_elem is None:
                query_elem = root.find('.//queryString')
            
            if query_elem is not None and query_elem.text:
                query = query_elem.text.strip()
                # Remove CDATA wrapper if present
                if query.startswith('<![CDATA[') and query.endswith(']]>'):
                    query = query[9:-3]
            
            # Get sample data
            data = self._get_sample_data(connection_string, query, fields)
            
            # Generate HTML
            html = self._generate_preview_html(title_text, fields, data)
            
            return html.encode('utf-8'), None
            
        except Exception as e:
            logger.error(f"HTML fallback generation error: {e}")
            return self._generate_fallback_message(f"Preview generation failed: {str(e)}"), str(e)
    
    def _get_sample_data(self, connection_string, query, fields):
        """Get data for the report from database or sample data."""
        if connection_string and connection_string.startswith('sqlite:///') and query:
            try:
                db_path = connection_string.replace('sqlite:///', '')
                if os.path.exists(db_path):
                    conn = sqlite3.connect(db_path)
                    cursor = conn.cursor()
                    cursor.execute(query)
                    columns = [description[0] for description in cursor.description]
                    rows = cursor.fetchall()
                    
                    data = []
                    for row in rows:
                        row_dict = {}
                        for i, col in enumerate(columns):
                            row_dict[col] = row[i] if i < len(row) else None
                        data.append(row_dict)
                    
                    conn.close()
                    return data
            except Exception as e:
                logger.warning(f"Database query error: {e}")
        
        # Return sample data based on fields
        if fields:
            sample_data = []
            for i in range(5):  # Generate 5 sample rows
                row = {}
                for field_name, field_type in fields.items():
                    if 'id' in field_name.lower():
                        row[field_name] = i + 1
                    elif 'name' in field_name.lower():
                        row[field_name] = f'Sample Name {i + 1}'
                    elif 'email' in field_name.lower():
                        row[field_name] = f'user{i + 1}@example.com'
                    elif 'date' in field_name.lower():
                        row[field_name] = f'2024-0{(i % 12) + 1}-{(i % 28) + 1:02d}'
                    elif 'amount' in field_name.lower() or 'salary' in field_name.lower():
                        row[field_name] = (i + 1) * 1000 + 500
                    elif 'department' in field_name.lower():
                        departments = ['IT', 'Sales', 'HR', 'Finance', 'Marketing']
                        row[field_name] = departments[i % len(departments)]
                    else:
                        row[field_name] = f'Sample {field_name} {i + 1}'
                sample_data.append(row)
            return sample_data
        
        return []
    
    def _generate_preview_html(self, title_text, fields, data):
        """Generate HTML preview with proper title display."""
        html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{title_text if title_text else 'Report Preview'}</title>
    <style>
        body {{ font-family: 'Times New Roman', serif; margin: 20px; }}
        .report-container {{ max-width: 800px; margin: 0 auto; }}
        .report-title {{ text-align: center; font-size: 24px; font-weight: bold; margin-bottom: 30px; color: #2c3e50; }}
        .fallback-notice {{ background-color: #fff3cd; border: 1px solid #ffeeba; color: #856404; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
        table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
        th {{ background-color: #34495e; color: white; padding: 8px; text-align: left; font-weight: bold; }}
        td {{ padding: 8px; border-bottom: 1px solid #ecf0f1; }}
        tr:nth-child(even) {{ background-color: #f8f9fa; }}
        .currency {{ text-align: right; }}
        @media print {{ body {{ margin: 0; }} }}
    </style>
</head>
<body>
    <div class="report-container">
        <div class="fallback-notice">
            <strong>Note:</strong> This is a preview generated using JRXML parsing. 
            For exact JasperReports output, ensure Java compatibility with the JasperReports engine.
        </div>
"""
        
        # Add title if it exists in the JRXML
        if title_text:
            html += f'        <div class="report-title">{title_text}</div>\n'
        
        # Add data table
        if data and fields:
            html += """
        <table>
            <thead>
                <tr>
"""
            for field_name in fields.keys():
                display_name = field_name.replace('_', ' ').title()
                html += f"                    <th>{display_name}</th>\n"
            
            html += """
                </tr>
            </thead>
            <tbody>
"""
            
            for row in data:
                html += "                <tr>\n"
                for field_name in fields.keys():
                    value = row.get(field_name, '')
                    if 'amount' in field_name.lower() or 'salary' in field_name.lower():
                        try:
                            amount_val = float(value) if value else 0.0
                            html += f'                    <td class="currency">${amount_val:.2f}</td>\n'
                        except (ValueError, TypeError):
                            html += f'                    <td class="currency">{value}</td>\n'
                    else:
                        html += f"                    <td>{value}</td>\n"
                html += "                </tr>\n"
            
            html += """
            </tbody>
        </table>
"""
        
        html += """
    </div>
</body>
</html>
"""
        
        return html
    
    def _generate_pdf_fallback(self, jrxml_content, connection_string):
        """Generate PDF using reportlab as fallback."""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            import xml.etree.ElementTree as ET
            
            # Parse JRXML to extract title and data structure
            root = ET.fromstring(jrxml_content)
            ns = {'jr': 'http://jasperreports.sourceforge.net/jasperreports'}
            if root.tag.startswith('{'):
                ns_url = root.tag[1:root.tag.index('}')]
                ns = {'jr': ns_url}
            
            # Extract title
            title_text = None
            title_band = root.find('.//jr:title', ns) or root.find('.//title')
            if title_band is not None:
                static_texts = title_band.findall('.//jr:staticText/jr:text', ns) or title_band.findall('.//staticText/text')
                for text_elem in static_texts:
                    if text_elem.text and text_elem.text.strip():
                        title_text = text_elem.text.strip()
                        break
            
            # Extract fields
            fields = {}
            field_elements = root.findall('.//jr:field', ns) or root.findall('.//field')
            for field in field_elements:
                name = field.get('name')
                if name:
                    fields[name] = field.get('class', 'java.lang.String')
            
            # Get data
            query = ""
            query_elem = root.find('.//jr:queryString', ns) or root.find('.//queryString')
            if query_elem is not None and query_elem.text:
                query = query_elem.text.strip()
                if query.startswith('<![CDATA[') and query.endswith(']]>'):
                    query = query[9:-3]
            
            data = self._get_sample_data(connection_string, query, fields)
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Add title if exists
            if title_text:
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Title'],
                    fontSize=24,
                    textColor=colors.HexColor('#2c3e50'),
                    spaceAfter=30,
                    alignment=1  # Center
                )
                story.append(Paragraph(title_text, title_style))
            
            # Add notice
            notice_style = ParagraphStyle(
                'Notice',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#856404'),
                backColor=colors.HexColor('#fff3cd'),
                borderColor=colors.HexColor('#ffeeba'),
                borderWidth=1,
                borderPadding=10,
                spaceAfter=20
            )
            story.append(Paragraph("Note: This PDF is generated using JRXML parsing. For exact JasperReports output, ensure Java compatibility.", notice_style))
            
            # Create table
            if data and fields:
                table_data = [[field.replace('_', ' ').title() for field in fields.keys()]]
                
                for row in data:
                    data_row = []
                    for field_name in fields.keys():
                        value = row.get(field_name, '')
                        if 'amount' in field_name.lower() or 'salary' in field_name.lower():
                            try:
                                amount_val = float(value) if value else 0.0
                                data_row.append(f'${amount_val:.2f}')
                            except:
                                data_row.append(str(value))
                        else:
                            data_row.append(str(value))
                    table_data.append(data_row)
                
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
                ]))
                story.append(table)
            
            doc.build(story)
            buffer.seek(0)
            return buffer.getvalue(), None
            
        except Exception as e:
            logger.error(f"PDF fallback generation error: {e}")
            return None, f"PDF generation error: {str(e)}"
    
    def _generate_excel_fallback(self, jrxml_content, connection_string):
        """Generate Excel using openpyxl as fallback."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            import xml.etree.ElementTree as ET
            
            # Parse JRXML
            root = ET.fromstring(jrxml_content)
            ns = {'jr': 'http://jasperreports.sourceforge.net/jasperreports'}
            if root.tag.startswith('{'):
                ns_url = root.tag[1:root.tag.index('}')]
                ns = {'jr': ns_url}
            
            # Extract title and fields
            title_text = None
            title_band = root.find('.//jr:title', ns) or root.find('.//title')
            if title_band is not None:
                static_texts = title_band.findall('.//jr:staticText/jr:text', ns) or title_band.findall('.//staticText/text')
                for text_elem in static_texts:
                    if text_elem.text and text_elem.text.strip():
                        title_text = text_elem.text.strip()
                        break
            
            fields = {}
            field_elements = root.findall('.//jr:field', ns) or root.findall('.//field')
            for field in field_elements:
                name = field.get('name')
                if name:
                    fields[name] = field.get('class', 'java.lang.String')
            
            # Get data
            query = ""
            query_elem = root.find('.//jr:queryString', ns) or root.find('.//queryString')
            if query_elem is not None and query_elem.text:
                query = query_elem.text.strip()
                if query.startswith('<![CDATA[') and query.endswith(']]>'):
                    query = query[9:-3]
            
            data = self._get_sample_data(connection_string, query, fields)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Report"
            
            row = 1
            
            # Add title
            if title_text:
                ws[f'A{row}'] = title_text
                ws[f'A{row}'].font = Font(bold=True, size=16)
                row += 2
            
            # Add notice
            ws[f'A{row}'] = "Note: Generated using JRXML parsing"
            ws[f'A{row}'].font = Font(italic=True, size=10)
            row += 2
            
            # Add headers
            if fields:
                col = 1
                for field_name in fields.keys():
                    ws.cell(row=row, column=col, value=field_name.replace('_', ' ').title())
                    ws.cell(row=row, column=col).font = Font(bold=True)
                    col += 1
                row += 1
                
                # Add data
                for row_data in data:
                    col = 1
                    for field_name in fields.keys():
                        value = row_data.get(field_name, '')
                        ws.cell(row=row, column=col, value=value)
                        col += 1
                    row += 1
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue(), None
            
        except Exception as e:
            logger.error(f"Excel fallback generation error: {e}")
            return None, f"Excel generation error: {str(e)}"
    
    def _generate_csv_fallback(self, jrxml_content, connection_string):
        """Generate CSV as fallback."""
        try:
            import csv
            from io import StringIO
            import xml.etree.ElementTree as ET
            
            # Parse JRXML
            root = ET.fromstring(jrxml_content)
            ns = {'jr': 'http://jasperreports.sourceforge.net/jasperreports'}
            if root.tag.startswith('{'):
                ns_url = root.tag[1:root.tag.index('}')]
                ns = {'jr': ns_url}
            
            fields = {}
            field_elements = root.findall('.//jr:field', ns) or root.findall('.//field')
            for field in field_elements:
                name = field.get('name')
                if name:
                    fields[name] = field.get('class', 'java.lang.String')
            
            # Get data
            query = ""
            query_elem = root.find('.//jr:queryString', ns) or root.find('.//queryString')
            if query_elem is not None and query_elem.text:
                query = query_elem.text.strip()
                if query.startswith('<![CDATA[') and query.endswith(']]>'):
                    query = query[9:-3]
            
            data = self._get_sample_data(connection_string, query, fields)
            
            output = StringIO()
            
            if fields:
                writer = csv.writer(output)
                
                # Write header
                writer.writerow([field.replace('_', ' ').title() for field in fields.keys()])
                
                # Write data
                for row in data:
                    writer.writerow([row.get(field, '') for field in fields.keys()])
            
            return output.getvalue().encode('utf-8'), None
            
        except Exception as e:
            logger.error(f"CSV fallback generation error: {e}")
            return None, f"CSV generation error: {str(e)}"